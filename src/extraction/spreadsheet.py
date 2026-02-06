"""Spreadsheet text extraction."""

import csv
import io
from pathlib import Path


def extract_text_from_spreadsheet(path: str | Path) -> tuple[str, dict]:
    """Extract text from a spreadsheet file.

    Returns (text, metadata) where text is tab-separated rows
    with sheet headers, and metadata contains sheet_count and row_count.
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return _extract_xlsx(path)
    elif ext == ".xls":
        return _extract_xls(path)
    elif ext == ".csv":
        return _extract_csv(path)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {ext}")


def _extract_xlsx(path: Path) -> tuple[str, dict]:
    """Extract text from .xlsx using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):  # skip fully empty rows
                lines.append("\t".join(cells))
                total_rows += 1

    wb.close()

    text = "\n".join(lines)
    metadata = {"sheet_count": len(wb.sheetnames), "row_count": total_rows}
    return text, metadata


def _extract_xls(path: Path) -> tuple[str, dict]:
    """Extract text from legacy .xls using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(str(path))
    lines = []
    total_rows = 0

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        lines.append(f"--- Sheet: {ws.name} ---")
        for row_idx in range(ws.nrows):
            cells = [str(ws.cell_value(row_idx, col)) for col in range(ws.ncols)]
            if any(cells):
                lines.append("\t".join(cells))
                total_rows += 1

    text = "\n".join(lines)
    metadata = {"sheet_count": wb.nsheets, "row_count": total_rows}
    return text, metadata


def _extract_csv(path: Path) -> tuple[str, dict]:
    """Extract text from CSV."""
    lines = []
    total_rows = 0

    # Try common encodings
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if any(row):
                        lines.append("\t".join(row))
                        total_rows += 1
            break
        except (UnicodeDecodeError, UnicodeError):
            lines.clear()
            total_rows = 0
            continue

    text = "\n".join(lines)
    metadata = {"sheet_count": 1, "row_count": total_rows}
    return text, metadata
